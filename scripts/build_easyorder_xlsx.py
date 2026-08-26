#!/usr/bin/env python3
"""Build EasyOrder-compatible Excel file from scraped Ecomerg products.

Implements:
- All 2638 products in the EasyOrder import format (same columns as sample)
- +1 KWD markup on every sale_price (price column = original + 2, sale_price = original + 1)
  so the customer sees a 1 KWD visual discount and the user earns 1 KWD markup.
- Color/size variations for the ~170 products that have them.
- Best sellers as a featured sheet.
- Categories sheet for reference.
- Verifies all image URLs return HTTP 200 (logs any broken ones).
"""

import json
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.request import urlopen, Request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------- Configuration ----------
PRODUCTS_JSON = "/home/z/my-project/download/ecomerg_products.json"
META_JSON = "/home/z/my-project/download/ecomerg_meta.json"
OUTPUT_XLSX = "/home/z/my-project/download/ecomerg_easyorder_import.xlsx"
BROKEN_LOG = "/home/z/my-project/download/broken_images.txt"

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

# Arabic color name → hex code mapping (for EasyOrder color variation)
COLOR_HEX = {
    "أسود": "#000000",
    "أبيض": "#FFFFFF",
    "بيج": "#F5F5DC",
    "أزرق": "#0000FF",
    "أحمر": "#FF0000",
    "أخضر": "#008000",
    "بني": "#8B4513",
    "برتقالي": "#FFA500",
    "وردي": "#FFC0CB",
    "رمادي": "#808080",
    "كحلي": "#1f3a5f",
    "ذهبي": "#d4af37",
    "فضي": "#C0C0C0",
    "أصفر": "#FFFF00",
    "بنفسجي": "#800080",
    "Natural": "#DEB887",
    "lvory": "#FFFFF0",
    "ivory": "#FFFFF0",
    "sky": "#87CEEB",
    "navy": "#1f3a5f",
    "brown": "#8B4513",
    "black": "#000000",
    "white": "#FFFFFF",
    "red": "#FF0000",
    "blue": "#0000FF",
    "green": "#008000",
    "yellow": "#FFFF00",
    "purple": "#800080",
    "pink": "#FFC0CB",
    "gray": "#808080",
    "grey": "#808080",
    "orange": "#FFA500",
    "gold": "#d4af37",
    "silver": "#C0C0C0",
    "multicolor": "#FF00FF",
}


def slugify(text):
    """Make a URL-safe slug from product code or name."""
    if not text:
        return ""
    # If text is a code (alphanumeric + dash), use as-is
    if re.match(r"^[a-zA-Z0-9\-]+$", text):
        return text.lower()
    # Transliterate Arabic to placeholders — fall back to a generic code
    # For Arabic-only strings, use a generated slug
    slug = re.sub(r"[^\w\s\-]", "", text, flags=re.UNICODE)
    slug = re.sub(r"[\s\-]+", "-", slug).strip("-").lower()
    return slug[:60] if slug else ""


def make_meta_description(name, description, max_len=160):
    """Build a short SEO meta description."""
    # Prefer description (cleaned), fallback to name
    desc = (description or "").strip()
    if not desc:
        return name or ""
    # Strip whitespace/newlines
    desc = re.sub(r"\s+", " ", desc)
    if len(desc) <= max_len:
        return desc
    # Truncate at word boundary
    truncated = desc[:max_len]
    last_space = truncated.rfind(" ")
    if last_space > 80:
        return truncated[:last_space] + "…"
    return truncated + "…"


def build_color_variation(colors_text):
    """Build variation string for color: 'اللون(color): name1=#hex1, name2=#hex2'."""
    if not colors_text or colors_text == "بدون لون":
        return None
    parts = []
    for name in re.split(r"[,،]", colors_text):
        name = name.strip()
        if not name or name == "بدون لون":
            continue
        hex_code = COLOR_HEX.get(name)
        if hex_code:
            parts.append(f"{name}={hex_code}")
        else:
            # Unknown color — keep the name without hex (still valid)
            parts.append(f"{name}=#CCCCCC")
    if not parts:
        return None
    return f"اللون(color): {', '.join(parts)}"


def build_size_variation(sizes_text):
    """Build variation string for size: 'المقاس(dropdown): S, M, L, XL'."""
    if not sizes_text or sizes_text == "بدون مقاس":
        return None
    # Split by comma, filter out "بدون مقاس"
    sizes = [s.strip() for s in re.split(r"[,،]", sizes_text) if s.strip() and s.strip() != "بدون مقاس"]
    if not sizes:
        return None
    return f"المقاس(dropdown): {', '.join(sizes)}"


def verify_image_url(url, timeout=15):
    """Verify an image URL returns HTTP 200 with image content type."""
    try:
        req = Request(url, method="HEAD", headers={
            "User-Agent": USER_AGENT,
            "Accept": "image/*,*/*;q=0.8",
        })
        with urlopen(req, timeout=timeout) as resp:
            if resp.status == 200:
                ctype = resp.headers.get("Content-Type", "")
                if "image" in ctype:
                    return True, "OK"
                return False, f"not image: {ctype}"
            return False, f"HTTP {resp.status}"
    except Exception as e:
        return False, str(e)[:80]


def verify_images_concurrently(urls, workers=20):
    """Verify many image URLs in parallel; return set of broken URLs."""
    broken = {}
    unique_urls = list(set(urls))
    print(f"Verifying {len(unique_urls)} unique image URLs...", flush=True)
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(verify_image_url, u): u for u in unique_urls}
        done = 0
        for fut in as_completed(futures):
            url = futures[fut]
            ok, msg = fut.result()
            if not ok:
                broken[url] = msg
            done += 1
            if done % 200 == 0:
                print(f"  Progress: {done}/{len(unique_urls)} (broken: {len(broken)})", flush=True)
    print(f"  Done. Broken: {len(broken)} / {len(unique_urls)}", flush=True)
    return broken


def main():
    # Load data
    with open(PRODUCTS_JSON, encoding="utf-8") as f:
        products_data = json.load(f)
    with open(META_JSON, encoding="utf-8") as f:
        meta = json.load(f)

    products = products_data["products"]
    best_seller_ids = {bs["id"] for bs in meta["best_sellers"]}
    categories = meta["categories"]

    # 1) Verify image URLs (sample first 50 for speed, then verify all in background)
    all_image_urls = [p["image_url"] for p in products if p.get("image_url")]
    broken = verify_images_concurrently(all_image_urls, workers=30)
    if broken:
        with open(BROKEN_LOG, "w", encoding="utf-8") as f:
            for url, msg in broken.items():
                f.write(f"{url}\t{msg}\n")
        print(f"  Broken URLs logged to {BROKEN_LOG}", flush=True)

    # 2) Build the workbook
    wb = Workbook()
    wb.properties.creator = "Z.ai"

    # ---- Sheet 1: Products (EasyOrder import format) ----
    ws = wb.active
    ws.title = "Products"

    headers = [
        "slug", "name", "price", "sale_price", "sku", "categories",
        "thumb", "images", "quantity", "track_stock",
        "disable_orders_for_no_stock", "description", "meta_description",
        "variation1", "variation2", "variation3", "variation4", "variation5", "variation6",
    ]
    ws.append(headers)

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Build parent-category map for subcategory lookup
    cat_map = {}
    for c in categories:
        if not c.get("is_subcategory"):
            cat_map[c["id"]] = c["name"]

    # Track missing categories
    missing_cats = 0

    for p in products:
        # Original price in KWD
        try:
            original_price = float(p.get("sale_price", {}).get("value", 0))
        except (TypeError, ValueError):
            original_price = 0

        # Apply user's +1 KWD markup:
        #   - sale_price (actual selling price) = original_price + 1
        #   - price (list/discounted-shown price) = original_price + 2  (1 KWD visual discount)
        sale_price = original_price + 1
        list_price = original_price + 2

        # Categories: main, sub (EasyOrder expects comma-separated)
        cats = []
        main_cat = (p.get("main_category") or "").strip()
        sub_cat = (p.get("sub_category") or "").strip()
        if main_cat:
            cats.append(main_cat)
        if sub_cat:
            cats.append(sub_cat)
        categories_str = ", ".join(cats)

        # Description
        desc = (p.get("description") or "").strip()
        # Replace any HTML entities that might exist
        desc = desc.replace("&nbsp;", " ").replace("&amp;", "&")

        # Meta description
        meta_desc = make_meta_description(p.get("name"), desc)

        # Variations
        variations = []
        color_var = build_color_variation(p.get("colors"))
        size_var = build_size_variation(p.get("sizes"))
        if color_var:
            variations.append(color_var)
        if size_var:
            variations.append(size_var)
        # Pad to 6 variation columns
        while len(variations) < 6:
            variations.append("")

        # Stock: best sellers get higher quantity, default 1000 for all
        quantity = 5000 if p["id"] in best_seller_ids else 1000

        # Build the row
        row = [
            slugify(p.get("code")) or f"product-{p['id']}",  # slug
            p.get("name", ""),                                # name
            list_price,                                       # price (list)
            sale_price,                                       # sale_price (selling)
            p.get("code", ""),                                # sku
            categories_str,                                   # categories
            p.get("image_url", ""),                           # thumb
            p.get("image_url", ""),                           # images (we only have one per product)
            quantity,                                         # quantity
            False,                                            # track_stock
            False,                                            # disable_orders_for_no_stock
            desc,                                             # description
            meta_desc,                                        # meta_description
            *variations,                                      # variation1..6
        ]
        ws.append(row)

    # Column widths
    col_widths = {
        "A": 18, "B": 40, "C": 8, "D": 10, "E": 14, "F": 25,
        "G": 60, "H": 60, "I": 10, "J": 12, "K": 18, "L": 50,
        "M": 35, "N": 35, "O": 35, "P": 35, "Q": 35, "R": 35, "S": 35,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Freeze the header row
    ws.freeze_panes = "A2"

    print(f"Products sheet: {ws.max_row - 1} products", flush=True)

    # ---- Sheet 2: Best Sellers ----
    ws2 = wb.create_sheet("Best Sellers")
    bs_headers = ["rank", "id", "code", "name", "sale_price_kwd", "commission_kwd", "image_url", "product_url"]
    ws2.append(bs_headers)
    for col_idx, _ in enumerate(bs_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    for rank, bs in enumerate(meta["best_sellers"], start=1):
        try:
            price_val = float(re.search(r"(\d+(?:\.\d+)?)", bs.get("sale_price") or "").group(1))
        except (AttributeError, ValueError):
            price_val = None
        try:
            comm_val = float(re.search(r"(\d+(?:\.\d+)?)", bs.get("commission") or "").group(1))
        except (AttributeError, ValueError):
            comm_val = None
        ws2.append([
            rank,
            bs.get("id"),
            bs.get("code"),
            bs.get("name", ""),
            price_val,
            comm_val,
            bs.get("image_url", ""),
            bs.get("product_url", ""),
        ])
    for col, w in {"A": 6, "B": 10, "C": 14, "D": 50, "E": 15, "F": 15, "G": 60, "H": 45}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"
    print(f"Best Sellers sheet: {ws2.max_row - 1} products", flush=True)

    # ---- Sheet 3: Categories ----
    ws3 = wb.create_sheet("Categories")
    cat_headers = ["category_id", "sub_id", "name", "is_subcategory", "products_url"]
    ws3.append(cat_headers)
    for col_idx, _ in enumerate(cat_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22

    for c in categories:
        ws3.append([
            c.get("id", ""),
            c.get("sub_id", ""),
            c.get("name", ""),
            "Yes" if c.get("is_subcategory") else "No",
            c.get("url", ""),
        ])
    for col, w in {"A": 12, "B": 10, "C": 35, "D": 15, "E": 50}.items():
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"
    print(f"Categories sheet: {ws3.max_row - 1} categories", flush=True)

    # ---- Sheet 4: Notes ----
    ws4 = wb.create_sheet("Import Notes")
    notes = [
        ["ملف الاستيراد لمتجرك على EasyOrder", ""],
        ["المصدر", "https://ecomerg.com"],
        ["عدد المنتجات", str(len(products))],
        ["عدد المنتجات الأكثر مبيعاً", str(len(meta["best_sellers"]))],
        ["عدد الفئات", str(len(categories))],
        ["", ""],
        ["شرح الأسعار", ""],
        ["السعر الأصلي للموقع (KWD)", "السعر المعروض على Ecomerg"],
        ["sale_price في الملف (KWD)", "السعر الأصلي + 1 دينار (ربح المستخدم)"],
        ["price في الملف (KWD)", "السعر الأصلي + 2 دينار (لإظهار خصم 1 دينار)"],
        ["مثال: منتج سعره 6 KWD", "price=8, sale_price=7  (المستخدم يربح 1 دينار والعميل يرى خصم 1 دينار)"],
        ["", ""],
        ["الحقول المعبأة", ""],
        ["slug", "كود المنتج من Ecomerg (مثل toy0075)"],
        ["name", "الاسم بالعربية من الموقع"],
        ["price / sale_price", "بعد إضافة الدينار كما هو موضح أعلاه"],
        ["sku", "نفس كود المنتج من Ecomerg"],
        ["categories", "القسم الرئيسي + القسم الفرعي (مفصولين بفاصلة)"],
        ["thumb / images", "رابط الصورة الكامل من موقع Ecomerg (تم التحقق منه)"],
        ["quantity", "1000 افتراضياً، 5000 للمنتججات الأكثر مبيعاً"],
        ["track_stock / disable_orders", "False (لا توقف الطلب عند نفاد المخزون)"],
        ["description", "الوصف الكامل بالعربية"],
        ["meta_description", "ملخص قصير للوصف (مناسب للسيو)"],
        ["variation1", "ألوان المنتج (إن وُجدت) بصيغة: اللون(color): أسود=#000000, أبيض=#FFFFFF"],
        ["variation2", "مقاسات المنتج (إن وُجدت) بصيغة: المقاس(dropdown): S, M, L, XL"],
        ["variation3..6", "فارغة (لا توجد اختيارات إضافية)"],
        ["", ""],
        ["المنتجات الأكثر مبيعاً", "راجع شيت Best Sellers لقائمة الـ 44 منتج"],
        ["الفئات", "راجع شيت Categories لقائمة الـ 38 فئة (رئيسية وفرعية)"],
        ["", ""],
        ["تنبيه", "تم فحص روابط الصور جميعها؛ أي صورة معطوبة مسجلة في broken_images.txt"],
    ]
    for row in notes:
        ws4.append(row)
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 80
    # Bold first column
    for row_idx in range(1, len(notes) + 1):
        ws4.cell(row=row_idx, column=1).font = Font(bold=True)
    print("Notes sheet created", flush=True)

    # Save
    wb.save(OUTPUT_XLSX)
    import os
    size_kb = os.path.getsize(OUTPUT_XLSX) / 1024
    print(f"\nSaved: {OUTPUT_XLSX} ({size_kb:.1f} KB)", flush=True)


if __name__ == "__main__":
    main()
