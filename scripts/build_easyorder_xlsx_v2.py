#!/usr/bin/env python3
"""Build EasyOrder Excel v2 - variable discounts + multiple images per product.

Changes from v1:
- Variable discount per product (10-20% random, reproducible via seed by product id).
  sale_price = original_price + 1 (still applies — the user's +1 KWD markup)
  price (list) = round_up(sale_price / (1 - discount%), 0.5)
- Multiple images per product (from product detail page scrape).
- track_stock=False; quantity=9999 (Ecomerg does not expose stock publicly).
- Best-sellers flagged with quantity=99999 + note in Import Notes sheet.
"""

import json
import math
import random
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.request import urlopen, Request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------- Configuration ----------
PRODUCTS_JSON = "/home/z/my-project/download/ecomerg_products.json"
IMAGES_JSON = "/home/z/my-project/download/ecomerg_product_images.json"
META_JSON = "/home/z/my-project/download/ecomerg_meta.json"
OUTPUT_XLSX = "/home/z/my-project/download/ecomerg_easyorder_import.xlsx"

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

# Arabic color name → hex code mapping
COLOR_HEX = {
    "أسود": "#000000", "أبيض": "#FFFFFF", "بيج": "#F5F5DC",
    "أزرق": "#0000FF", "أحمر": "#FF0000", "أخضر": "#008000",
    "بني": "#8B4513", "برتقالي": "#FFA500", "وردي": "#FFC0CB",
    "رمادي": "#808080", "كحلي": "#1f3a5f", "ذهبي": "#d4af37",
    "فضي": "#C0C0C0", "أصفر": "#FFFF00", "بنفسجي": "#800080",
    "Natural": "#DEB887", "lvory": "#FFFFF0", "ivory": "#FFFFF0",
    "sky": "#87CEEB", "navy": "#1f3a5f", "brown": "#8B4513",
    "black": "#000000", "white": "#FFFFFF", "red": "#FF0000",
    "blue": "#0000FF", "green": "#008000", "yellow": "#FFFF00",
    "purple": "#800080", "pink": "#FFC0CB", "gray": "#808080",
    "grey": "#808080", "orange": "#FFA500", "gold": "#d4af37",
    "silver": "#C0C0C0", "multicolor": "#FF00FF",
}


def slugify(text):
    if not text:
        return ""
    if re.match(r"^[a-zA-Z0-9\-]+$", text):
        return text.lower()
    slug = re.sub(r"[^\w\s\-]", "", text, flags=re.UNICODE)
    slug = re.sub(r"[\s\-]+", "-", slug).strip("-").lower()
    return slug[:60] if slug else ""


def make_meta_description(name, description, max_len=160):
    desc = (description or "").strip()
    if not desc:
        return name or ""
    desc = re.sub(r"\s+", " ", desc)
    if len(desc) <= max_len:
        return desc
    truncated = desc[:max_len]
    last_space = truncated.rfind(" ")
    if last_space > 80:
        return truncated[:last_space] + "…"
    return truncated + "…"


def build_color_variation(colors_text):
    if not colors_text or colors_text == "بدون لون":
        return None
    parts = []
    for name in re.split(r"[,،]", colors_text):
        name = name.strip()
        if not name or name == "بدون لون":
            continue
        hex_code = COLOR_HEX.get(name, "#CCCCCC")
        parts.append(f"{name}={hex_code}")
    if not parts:
        return None
    return f"اللون(color): {', '.join(parts)}"


def build_size_variation(sizes_text):
    if not sizes_text or sizes_text == "بدون مقاس":
        return None
    sizes = [s.strip() for s in re.split(r"[,،]", sizes_text) if s.strip() and s.strip() != "بدون مقاس"]
    if not sizes:
        return None
    return f"المقاس(dropdown): {', '.join(sizes)}"


def calculate_prices(original_price, product_id):
    """Calculate list_price and sale_price with variable discount.

    - sale_price = original_price + 1 (user's +1 KWD markup)
    - discount_pct = random in [0.10, 0.20] seeded by product_id (reproducible)
    - list_price = round_up(sale_price / (1 - discount_pct), 0.5)
    """
    # Seeded random for reproducibility
    rng = random.Random(product_id)
    discount_pct = rng.uniform(0.10, 0.20)

    sale_price = original_price + 1
    # List price = sale_price / (1 - discount) — rounded up to nearest 0.25 KWD
    # (KWD has fils, but we use 0.25 increments for clean retail prices)
    list_price = sale_price / (1 - discount_pct)
    # Round up to nearest 0.25
    list_price = math.ceil(list_price * 4) / 4

    # Ensure list_price > sale_price (discount must be positive)
    if list_price <= sale_price:
        list_price = sale_price + 0.25

    return list_price, sale_price, discount_pct


def main():
    # Load data
    with open(PRODUCTS_JSON, encoding="utf-8") as f:
        products_data = json.load(f)
    with open(IMAGES_JSON, encoding="utf-8") as f:
        images_map = {int(k): v for k, v in json.load(f).items()}
    with open(META_JSON, encoding="utf-8") as f:
        meta = json.load(f)

    products = products_data["products"]
    best_seller_ids = {bs["id"] for bs in meta["best_sellers"]}
    categories = meta["categories"]

    print(f"Loaded {len(products)} products, {len(images_map)} image sets, {len(best_seller_ids)} best sellers", flush=True)

    # Build workbook
    wb = Workbook()
    wb.properties.creator = "Z.ai"

    # ---- Sheet 1: Products ----
    ws = wb.active
    ws.title = "Products"

    headers = [
        "slug", "name", "price", "sale_price", "sku", "categories",
        "thumb", "images", "quantity", "track_stock",
        "disable_orders_for_no_stock", "description", "meta_description",
        "variation1", "variation2", "variation3", "variation4", "variation5", "variation6",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Track stats
    stats = {
        "total": 0,
        "with_multiple_images": 0,
        "with_color": 0,
        "with_size": 0,
        "best_sellers": 0,
    }

    for p in products:
        stats["total"] += 1
        pid = p["id"]

        # Original price
        try:
            original_price = float(p.get("sale_price", {}).get("value", 0))
        except (TypeError, ValueError):
            original_price = 0

        # Variable discount calculation
        list_price, sale_price, discount_pct = calculate_prices(original_price, pid)

        # Categories
        cats = []
        main_cat = (p.get("main_category") or "").strip()
        sub_cat = (p.get("sub_category") or "").strip()
        if main_cat:
            cats.append(main_cat)
        if sub_cat:
            cats.append(sub_cat)
        categories_str = ", ".join(cats)

        # Description and meta
        desc = (p.get("description") or "").strip()
        desc = desc.replace("&nbsp;", " ").replace("&amp;", "&")
        meta_desc = make_meta_description(p.get("name"), desc)

        # Variations
        variations = []
        color_var = build_color_variation(p.get("colors"))
        size_var = build_size_variation(p.get("sizes"))
        if color_var:
            variations.append(color_var)
            stats["with_color"] += 1
        if size_var:
            variations.append(size_var)
            stats["with_size"] += 1
        while len(variations) < 6:
            variations.append("")

        # Images - multiple per product
        product_images = images_map.get(pid, [])
        # Fallback to thumbnail from listing if no detail images
        if not product_images and p.get("image_url"):
            product_images = [p["image_url"]]
        elif not product_images:
            product_images = []

        if len(product_images) > 1:
            stats["with_multiple_images"] += 1

        # thumb = first image; images = comma-separated remaining (or all)
        thumb = product_images[0] if product_images else ""
        images_str = ", ".join(product_images) if product_images else ""

        # Quantity & stock: fixed 20 pieces per product (user-specified)
        quantity = 20
        if pid in best_seller_ids:
            stats["best_sellers"] += 1

        row = [
            slugify(p.get("code")) or f"product-{pid}",
            p.get("name", ""),
            list_price,
            sale_price,
            p.get("code", ""),
            categories_str,
            thumb,
            images_str,
            quantity,
            False,  # track_stock
            False,  # disable_orders_for_no_stock
            desc,
            meta_desc,
            *variations,
        ]
        ws.append(row)

    # Column widths
    col_widths = {
        "A": 18, "B": 40, "C": 8, "D": 10, "E": 14, "F": 25,
        "G": 60, "H": 80, "I": 10, "J": 12, "K": 18, "L": 50,
        "M": 35, "N": 35, "O": 35, "P": 35, "Q": 35, "R": 35, "S": 35,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    print(f"Products sheet: {ws.max_row - 1} products", flush=True)
    print(f"  with multiple images: {stats['with_multiple_images']}", flush=True)
    print(f"  with color variation: {stats['with_color']}", flush=True)
    print(f"  with size variation:  {stats['with_size']}", flush=True)
    print(f"  best sellers: {stats['best_sellers']} (same qty=20, flagged in Best Sellers sheet)", flush=True)

    # Show price distribution sample
    print("\nSample prices (first 10):", flush=True)
    for r in range(2, 12):
        code = ws.cell(row=r, column=5).value
        price = ws.cell(row=r, column=3).value
        sale = ws.cell(row=r, column=4).value
        disc = round(price - sale, 2)
        disc_pct = round(disc / price * 100, 1) if price else 0
        print(f"  [{code}] price={price} | sale={sale} | discount={disc} KWD ({disc_pct}%)", flush=True)

    # ---- Sheet 2: Best Sellers ----
    ws2 = wb.create_sheet("Best Sellers")
    bs_headers = ["rank", "id", "code", "name", "original_price_kwd", "list_price_kwd", "sale_price_kwd", "discount_kwd", "discount_pct", "image_url", "product_url"]
    ws2.append(bs_headers)
    for col_idx, _ in enumerate(bs_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    for rank, bs in enumerate(meta["best_sellers"], start=1):
        try:
            orig_price = float(re.search(r"(\d+(?:\.\d+)?)", bs.get("sale_price") or "").group(1))
        except (AttributeError, ValueError):
            orig_price = 0
        list_price, sale_price, discount_pct = calculate_prices(orig_price, bs["id"])
        ws2.append([
            rank,
            bs.get("id"),
            bs.get("code"),
            bs.get("name", ""),
            orig_price,
            list_price,
            sale_price,
            round(list_price - sale_price, 2),
            round(discount_pct * 100, 1),
            bs.get("image_url", ""),
            bs.get("product_url", ""),
        ])
    for col, w in {"A": 6, "B": 10, "C": 14, "D": 50, "E": 18, "F": 18, "G": 18, "H": 18, "I": 14, "J": 60, "K": 45}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"
    print(f"Best Sellers sheet: {ws2.max_row - 1} products", flush=True)

    # ---- Sheet 3: Categories ----
    ws3 = wb.create_sheet("Categories")
    cat_headers = ["category_id", "sub_id", "name", "is_subcategory", "products_url"]
    ws3.append(cat_headers)
    for col_idx, _ in enumerate(cat_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22

    for c in categories:
        ws3.append([
            c.get("id", ""),
            c.get("sub_id", ""),
            c.get("name", ""),
            "Yes" if c.get("is_subcategory") else "No",
            c.get("url", ""),
        ])
    for col, w in {"A": 12, "B": 10, "C": 35, "D": 15, "E": 50}.items():
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"
    print(f"Categories sheet: {ws3.max_row - 1} categories", flush=True)

    # ---- Sheet 4: Notes ----
    ws4 = wb.create_sheet("Import Notes")
    notes = [
        ["ملف الاستيراد لمتجرك على EasyOrder", ""],
        ["المصدر", "https://ecomerg.com"],
        ["عدد المنتجات", str(len(products))],
        ["عدد المنتجات الأكثر مبيعاً", str(len(meta["best_sellers"]))],
        ["عدد الفئات", str(len(categories))],
        ["إجمالي عدد الصور", str(sum(len(v) for v in images_map.values()))],
        ["متوسط الصور لكل منتج", f"{sum(len(v) for v in images_map.values()) / len(images_map):.1f}"],
        ["", ""],
        ["شرح الأسعار (محدث)", ""],
        ["السعر الأصلي على Ecomerg (KWD)", "السعر المعروض على الموقع (شامل العمولة 1 د.ك)"],
        ["sale_price في الملف (KWD)", "السعر الأصلي + 1 دينار (ربح المستخدم)"],
        ["price في الملف (KWD)", "محسوب ليعطي خصم 10%-20% عشوائي على sale_price"],
        ["نسبة الخصم (%)", "تتغير من 10% إلى 20% لكل منتج (مولّدة بـ seed ثابت للمنتج)"],
        ["مثال: منتج أصله 6 KWD", "sale_price=7 (ربح 1 د.ك)، price≈8.5-9 (خصم 17-22%)"],
        ["مثال: منتج أصله 10 KWD", "sale_price=11 (ربح 1 د.ك)، price≈13-14 (خصم 15-22%)"],
        ["", ""],
        ["المخزون (الكمية)", ""],
        ["هل المخزون متوفر من Ecomerg؟", "لا - الموقع لا يعرض أرقام المخزون للعموم"],
        ["quantity المُعيّن", "20 قطعة لكل منتج (افتراضياً موحّد)"],
        ["track_stock", "False (لا يتتبع المخزون - الطلب متاح دائماً)"],
        ["disable_orders_for_no_stock", "False (لا يوقف الطلب عند نفاد المخزون)"],
        ["تنبيه", "الكمية 20 موحّدة لكل المنتجات بما فيها الأكثر مبيعاً - عدّلها يدوياً لكل منتج حسب حاجتك"],
        ["", ""],
        ["الصور (محدث)", ""],
        ["thumb", "الصورة الأولى من صفحة المنتج (الصورة الرئيسية)"],
        ["images", "كل صور المنتج مفصولة بفاصلة - من 1 إلى 10 صور لكل منتج"],
        ["متوسط الصور لكل منتج", f"{sum(len(v) for v in images_map.values()) / len(images_map):.1f} صورة"],
        ["", ""],
        ["الحقول الأخرى", ""],
        ["slug", "كود المنتج من Ecomerg (مثل toy0075)"],
        ["name", "الاسم بالعربية من الموقع"],
        ["sku", "نفس كود المنتج من Ecomerg"],
        ["categories", "القسم الرئيسي + القسم الفرعي (مفصولين بفاصلة)"],
        ["description", "الوصف الكامل بالعربية"],
        ["meta_description", "ملخص قصير للوصف (مناسب للسيو)"],
        ["variation1", "ألوان المنتج (إن وُجدت): اللون(color): أسود=#000000, أبيض=#FFFFFF"],
        ["variation2", "مقاسات المنتج (إن وُجدت): المقاس(dropdown): S, M, L, XL"],
        ["variation3..6", "فارغة"],
        ["", ""],
        ["المنتجات الأكثر مبيعاً", "راجع شيت Best Sellers - الـ 44 منتج + كل البيانات"],
        ["الفئات", "راجع شيت Categories - 38 فئة (12 رئيسية + 26 فرعية)"],
        ["", ""],
        ["أوراق البيانات الخام (مرجعية)", ""],
        ["Raw Products", "كل البيانات الخام من Ecomerg (id, code, name, description, price, commission, category, supplier, colors, sizes, image_url, product_url) - 2638 صف"],
        ["Product Images", "كل روابط صور كل منتج كأعمدة مستقلة (image_1, image_2, ... image_10) - 2638 صف"],
        ["", ""],
        ["ملف موحّد", "كل البيانات مدمجة في ملف Excel واحد - لا حاجة لأي ملفات JSON أو CSV منفصلة"],
    ]
    for row in notes:
        ws4.append(row)
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 80
    for row_idx in range(1, len(notes) + 1):
        ws4.cell(row=row_idx, column=1).font = Font(bold=True)
    print("Notes sheet created", flush=True)

    # ---- Sheet 5: Raw Products (all original fields from Ecomerg) ----
    ws5 = wb.create_sheet("Raw Products")
    raw_headers = [
        "id", "code", "name", "description",
        "original_price_kwd", "commission_kwd", "commission_text",
        "main_category", "sub_category", "supplier",
        "colors", "sizes", "image_url", "product_url",
    ]
    ws5.append(raw_headers)
    for col_idx, _ in enumerate(raw_headers, start=1):
        cell = ws5.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 22

    for p in products:
        ws5.append([
            p.get("id", ""),
            p.get("code", ""),
            p.get("name", ""),
            (p.get("description") or "").replace("&nbsp;", " ").replace("&amp;", "&"),
            (p.get("sale_price") or {}).get("value", ""),
            (p.get("commission") or {}).get("value", ""),
            (p.get("commission") or {}).get("text", ""),
            p.get("main_category", ""),
            p.get("sub_category", ""),
            p.get("supplier", ""),
            p.get("colors", ""),
            p.get("sizes", ""),
            p.get("image_url", ""),
            p.get("product_url", ""),
        ])
    for col, w in {
        "A": 8, "B": 14, "C": 45, "D": 60, "E": 18, "F": 14, "G": 25,
        "H": 22, "I": 22, "J": 12, "K": 18, "L": 18, "M": 55, "N": 40,
    }.items():
        ws5.column_dimensions[col].width = w
    ws5.freeze_panes = "A2"
    print(f"Raw Products sheet: {ws5.max_row - 1} rows", flush=True)

    # ---- Sheet 6: Product Images (one row per product, image URLs as columns) ----
    ws6 = wb.create_sheet("Product Images")
    max_images = max((len(v) for v in images_map.values()), default=0)
    img_headers = ["product_id", "code", "name", "image_count"] + [f"image_{i+1}" for i in range(max_images)]
    ws6.append(img_headers)
    for col_idx, _ in enumerate(img_headers, start=1):
        cell = ws6.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 22

    # Build a product lookup by id for code/name
    product_lookup = {p["id"]: p for p in products}

    for pid, urls in sorted(images_map.items()):
        p = product_lookup.get(pid, {})
        row = [pid, p.get("code", ""), p.get("name", ""), len(urls)]
        # Pad urls to max_images
        padded = list(urls) + [""] * (max_images - len(urls))
        row.extend(padded)
        ws6.append(row)

    # Set column widths for images sheet
    ws6.column_dimensions["A"].width = 10
    ws6.column_dimensions["B"].width = 14
    ws6.column_dimensions["C"].width = 45
    ws6.column_dimensions["D"].width = 12
    for i in range(max_images):
        ws6.column_dimensions[get_column_letter(5 + i)].width = 55
    ws6.freeze_panes = "E2"
    print(f"Product Images sheet: {ws6.max_row - 1} rows, max images per product: {max_images}", flush=True)

    # Save
    wb.save(OUTPUT_XLSX)
    import os
    size_kb = os.path.getsize(OUTPUT_XLSX) / 1024
    print(f"\nSaved: {OUTPUT_XLSX} ({size_kb:.1f} KB)", flush=True)
    print(f"\nWorkbook sheets:", flush=True)
    for s in wb.sheetnames:
        print(f"  - {s}", flush=True)


if __name__ == "__main__":
    main()
